import math, re, unicodedata, urllib.parse, os
from bs4 import BeautifulSoup
import requests
from collections import namedtuple
import random

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

INPUT_URL   = "https://www.mercadolibre.com.ar/ofertas"
OUTPUT_FILE = "articulos_mercado.xlsx"
DOWNLOAD_IMAGES = True

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36",
}

KEYWORDS = ["ps5", "playstation", "starlink", "rtx", "tv", "hisense", "samsung", "asus", "hp"]
KEYWORD_BOOST = 15.0

BRAND_BOOSTES = {
    "hp": 5.0, "asus": 5.0, "samsung": 7.0, "hisense": 3.0,
    "philips": 3.0, "philco": 2.0, "playstation": 10.0
}

WEIGHTS = {
    "pct_off": 0.30,
    "abs_saving": 0.25,
    "cheapness": 0.05,
    "keyword_brand": 0.40,
}

CAPS = {"pct_off": 80.0, "abs_saving": None}

def parse_money(s):
    """Convierte '1.234.567' o '123.456' en int"""
    if s is None:
        return None
    s = re.sub(r"[^\d]", "", str(s))
    return int(s) if s else None

def parse_pct_off(s):
    """Extrae número de '53% OFF' → 53.0"""
    if not s:
        return 0.0
    m = re.search(r"(\d+(?:[.,]\d+)?)", str(s))
    return float(m.group(1).replace(",", ".")) if m else 0.0

def score_item(it):
    """Calcula el score heurístico del producto"""
    precio_antes  = parse_money(it.get("precio_antes"))
    precio_actual = parse_money(it.get("precio_actual"))
    pct_off = parse_pct_off(it.get("descuento"))
    if CAPS["pct_off"] is not None:
        pct_off = min(pct_off, CAPS["pct_off"])

    abs_saving = 0.0
    if precio_antes and precio_actual and precio_antes > precio_actual:
        abs_saving = precio_antes - precio_actual

    cheapness = 0.0
    if precio_actual and precio_actual > 0:
        cheapness = 1.0 / math.log(precio_actual, 10)

    name = (it.get("nombre") or "").lower()
    kw_points = 0.0
    for kw in KEYWORDS:
        if kw in name:
            kw_points += KEYWORD_BOOST
    for brand, boost in BRAND_BOOSTES.items():
        if brand in name:
            kw_points += boost

    pct_norm   = pct_off / 100.0
    save_norm  = (abs_saving / 2_000_000.0) if abs_saving else 0.0
    cheap_norm = cheapness
    kw_norm    = min(1.0, kw_points / 30.0)

    final = (
        WEIGHTS["pct_off"]       * pct_norm +
        WEIGHTS["abs_saving"]    * save_norm +
        WEIGHTS["cheapness"]     * cheap_norm +
        WEIGHTS["keyword_brand"] * kw_norm
    )
    return final

ID_RE = re.compile(r"(MLA\d{6,})", re.IGNORECASE)

def product_ids_from_url(url: str):
    if not url:
        return set()
    ids = set(m.group(1).upper() for m in ID_RE.finditer(url))
    qs = urllib.parse.urlparse(url).query
    qd = urllib.parse.parse_qs(qs)
    for k in ("wid",):
        for v in qd.get(k, []):
            if isinstance(v, str) and v.upper().startswith("MLA"):
                ids.add(v.upper())
    return ids

def norm_name(name: str):
    """Normaliza nombre para detectar duplicados (agresivo pero simple)"""
    if not name:
        return ""
    name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    name = name.lower()
    STOP = [
        "color", "negro", "blanco", "gris", "azul", "rojo", "android", "google tv",
        "smart tv", "pantalla", "pulgadas", "full hd", "uhd", "4k", "8k", "wifi",
        "led", "qled", "ips", "mini", "kit", "combo"
    ]
    for w in STOP:
        name = name.replace(w, " ")
    name = re.sub(r"[^a-z0-9\s]", " ", name)
    name = re.sub(r"\b\d+\b", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name

def _pick_from_srcset(srcset: str):
    try:
        return srcset.split(",")[-1].strip().split()[0]
    except Exception:
        return None

def get_image_url(div_imagen):
    if not div_imagen:
        return None
    source = div_imagen.find("source")
    if source and source.get("srcset"):
        url = _pick_from_srcset(source["srcset"])
        if url:
            return url
    img = div_imagen.find("img")
    if not img:
        return None
    for attr in ("data-src", "data-original", "data-image", "data-lazy", "data-srcset"):
        if img.get(attr):
            if attr.endswith("srcset"):
                url = _pick_from_srcset(img[attr])
                if url:
                    return url
            else:
                return img[attr]
    if img.get("srcset"):
        url = _pick_from_srcset(img["srcset"])
        if url:
            return url
    src = img.get("src")
    if src and not src.startswith("data:"):
        return src
    return None

Articulo = namedtuple('Articulo', ['nombre', 'precio_antes', 'precio_actual', 'descuento',
                                   'link_a_comprar', 'imagen', 'img_id'])

def scrape_ofertas():
    r = requests.get(INPUT_URL, headers=HEADERS, timeout=30)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, 'html.parser')
    cards = soup.find_all('div', class_="poly-card")

    if DOWNLOAD_IMAGES:
        os.makedirs("./imagenes", exist_ok=True)

    items = []
    for card in cards:
        title_tag = card.find('h3', class_='poly-component__title-wrapper')
        if not title_tag:
            title_tag = card.find(['h2', 'h3'])
        nombre = title_tag.get_text(strip=True) if title_tag else None
        if not nombre:
            continue

        # Precios y descuento (tu lógica original)
        try:
            precio_antes_tag = card.find('span', class_="andes-money-amount__fraction")
            precio_antes = precio_antes_tag.get_text(strip=True) if precio_antes_tag else None
            descuento_tag = card.find('span', class_="andes-money-amount__discount")
            descuento = descuento_tag.get_text(strip=True) if descuento_tag else None
        except Exception:
            precio_antes, descuento = None, None

        precio_padre = card.find('div', class_="poly-price__current") or card
        precio_actual_tag = precio_padre.find('span', class_="andes-money-amount__fraction")
        precio_actual = precio_actual_tag.get_text(strip=True) if precio_actual_tag else None

        # Imagen (opcional)
        div_imagen = card.find('div', class_="poly-card__portada") or card.find('div', class_="poly-card__image")
        imagen = get_image_url(div_imagen)
        img_id = random.randint(10_000_000, 99_999_999)

        if DOWNLOAD_IMAGES and imagen:
            try:
                i = requests.get(imagen, headers=HEADERS, timeout=30)
                if i.ok:
                    with open(f'./imagenes/{img_id}.webp', 'wb') as f:
                        f.write(i.content)
            except Exception:
                pass

        # Link
        a = card.find('a')
        link_a_comprar = a['href'] if a and a.get('href') else None

        items.append(Articulo(
            nombre, precio_antes, precio_actual, descuento, link_a_comprar, imagen, img_id
        )._asdict())

    return items

EXCEL_COLUMNS = [
    ("nombre", "Nombre"),
    ("precio_antes", "Precio antes"),
    ("precio_actual", "Precio actual"),
    ("descuento", "Descuento"),
    ("link_a_comprar", "Link de compra"),
    ("_score", "Score relevancia"),
]

def write_excel(items, path=OUTPUT_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ofertas"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, (_, label) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"

    for r, it in enumerate(items, start=2):
        for col_idx, (key, _) in enumerate(EXCEL_COLUMNS, 1):
            val = it.get(key)
            cell = ws.cell(row=r, column=col_idx, value=val)

            if key == "link_a_comprar" and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.style = "Hyperlink"

    widths = {
        "Nombre": 55,
        "Precio antes": 14,
        "Precio actual": 14,
        "Descuento": 12,
        "Link de compra": 45,
        "Score relevancia": 16,
    }
    for i, (_, label) in enumerate(EXCEL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(label, 18)

    # Tabla + filtros
    last_row = ws.max_row
    last_col = len(EXCEL_COLUMNS)
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName="OfertasTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    wb.save(path)

def main():
    data = scrape_ofertas()

    for it in data:
        it["_score"] = round(score_item(it), 6)

    data_sorted = sorted(data, key=lambda x: x["_score"], reverse=True)

    seen_img_ids = set()
    seen_prod_ids = set()
    seen_names = set()
    deduped = []

    for it in data_sorted:
        img_id = it.get("img_id")
        if img_id in seen_img_ids:
            continue

        url = it.get("link_a_comprar", "") or it.get("link", "")
        pids = product_ids_from_url(url)
        if pids & seen_prod_ids:
            continue

        name_key = norm_name(it.get("nombre", ""))
        if name_key and name_key in seen_names:
            continue

        if img_id is not None:
            seen_img_ids.add(img_id)
        seen_prod_ids |= pids
        if name_key:
            seen_names.add(name_key)

        deduped.append(it)

    write_excel(deduped, OUTPUT_FILE)

    print(f"✅ Scrapeados: {len(data)} | Tras dedupe: {len(deduped)}")
    print(f"📄 Excel guardado en: {OUTPUT_FILE}")
    print("\nTOP 10 por relevancia:\n")
    for i, it in enumerate(deduped[:10], 1):
        print(f"{i:02d}. {it.get('nombre','')[:90]}...")
        print(f"    {it.get('descuento','')} | ${it.get('precio_actual','-')} | score={it['_score']}")

if __name__ == "__main__":
    main()
